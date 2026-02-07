from __future__ import annotations
from typing import List, Optional, Any, Dict, Tuple
from fastapi import UploadFile
from openpyxl import load_workbook

import io 
from app.services.upload.utils import build_header_map

def read_xlsx_rows(file: UploadFile) -> List[Dict[str, Any]]:
    """
    Read an XLSX UploadFile and extract raw rows containing only
    the required logical fields: platform and url.

    Input:
    - file: UploadFile pointing to an XLSX file.

    Output:
    - List of dicts, each with keys:
      {
        "platform": Any,
        "url": Any
      }

    Notes:
    - This function performs NO validation.
    - The header row is auto-detected by searching for
      columns named 'platform' and 'url'.
    - Empty rows are skipped.
    - Only the active worksheet is read.
    """

    file.file.seek(0)
    wb = load_workbook(filename=io.BytesIO(file.file.read()), read_only=True, data_only=True)
    ws = wb.active # read the first sheet
    
    header_row_idx, headers = _find_header_row(ws)
    if not headers:
        return [] # no header found, treat as empty
    
    key_map = build_header_map(headers)
    
    rows: List[Dict[str, Any]] = []
    for excel_row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
        if not excel_row or all(cell is None or str(cell).strip() == "" for cell in excel_row):
            continue # skip empty rows
        
        platform_val = _cell_by_header(excel_row, headers, key_map["platform"])
        url_val = _cell_by_header(excel_row, headers, key_map["url"])
        rows.append({
            "platform": platform_val,
            "url": url_val
        })
    return rows
    
    
def _find_header_row(ws) -> Tuple[int, List[str]]:
    """
    Locate the header row in an Excel worksheet.

    Strategy:
    - Scan rows from top to bottom.
    - The first row containing both 'platform' and 'url'
      (case-insensitive) is treated as the header.

    Returns:
    - (row_index, headers)

    Notes:
    - row_index is 1-based (Excel-style).
    - Falls back to the first row if no obvious header is found.
    """
    for idx,row in enumerate(ws.iter_rows(values_only=True),start = 1):
        if not row:
            continue
        headers = [str(cell).strip() if cell is not None else "" for cell in row]
        lowered = {h.lower() for h in headers if h}
        if "platform" in lowered and "url" in lowered:
            return idx, headers
    
    first = next(ws.iter_rows(values_only=True), None)
    if first:
        headers = [str(cell).strip() if cell is not None else "" for cell in first]
        return 1, headers
    return 1,[]


def _cell_by_header(excel_row: Tuple[Any, ...], headers: List[str], header_name: str) -> Any:
    """
    Retrieve a cell value from an Excel row by header name.

    Input:
    - excel_row: tuple of cell values
    - headers: list of header strings
    - header_name: exact header to locate

    Output:
    - Cell value if present, otherwise None.

    Notes:
    - No type conversion is performed here.
    """
    try:
        idx = headers.index(header_name)
    except ValueError:
        return None
    if idx >= len(excel_row) or idx < 0:
        return None
    return excel_row[idx]
    